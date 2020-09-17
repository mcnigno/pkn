def sequencer():  
    # main setting
    separator = "-"
    serial_digits = 4
    # projects table
    prj = {}
    # main codec elements table for prj
    prj_fields = {
        "unit": 3,
        "doctype": 3,
        "matclass": 3
    }
    # any element has many sub elements

    # extra codec information table for prj
    # this options will be added to the matrix.
    prj_extra_fields = {
        "partner": 3,
        "mr": 0,
    }

    # matrix table
    matrix = {}

from app.models import Matrix
from app import db
from flask import abort, Response, flash
def askcode(unit, discipline, doctype, quantity):
    #p = unit
    session = db.session
    codes_list = []
    
    project_number = "A8RX"
    entity = "TSC"
    
    params = [project_number, entity, str(unit.code), str(discipline.code), str(doctype.code)]
    print('ASKCODE  * * * * * * params', params)
    base_data = "-".join(params)
    
    for n in range(quantity):
        matrix = session.query(Matrix).filter(Matrix.base == base_data).first()

        if matrix:
            #matrix.created_by_fk = '1'
            matrix.changed_by_fk = '1'
            if discipline.stop:
                discipline_stop = discipline.stop
            else:
                discipline_stop = 999

            
            print('matrix counter', matrix.counter, 'discipline stop', discipline_stop)
            
            if matrix.counter + 1 <= discipline_stop:
                    matrix.counter += 1 
            else: 
                flash('Warning: Discipline '+str(discipline.name)+' -> Range Limit!', category='warning')
                abort(Response('Warning: Discipline Range Limit'))       
        else:
            matrix = Matrix()
            matrix.base = base_data
            
            if discipline.start:
                matrix.counter = discipline.start
            else:
                matrix.counter = 1
            matrix.created_by_fk = '1'
            matrix.changed_by_fk = '1'
            session.add(matrix)

        session.commit()
        codes_list.append(base_data + "-" + str(matrix.counter).zfill(3))
    return codes_list

from openpyxl import load_workbook
from app.models import Unit, Doctype, Discipline, Codes, Request
from app import db

def upload_old_codes():
    wb = load_workbook('app/static/uploads/pkn_old_codes.xlsx', read_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=5):
        print(row[1].value,row[2].value,row[3].value, row[5].value)
        discipline_code = row[1].value
        discipline_name = row[2].value
        unit_code = row[3].value
        released_code = row[4].value
        doctype = row[5].value
        contractor_code = row[7].value

        #add to internal note
        revision = 'Rev: ' + str(row[8].value)
        category = 'Cat: ' + str(row[9].value)
        submittal_purpose = 'Sub: '+ str(row[10].value)
        document_title = 'Title: '+str(row[11].value)
        due_date = 'DueDate: '+str(row[12].value)

        note = " ".join([revision, category, submittal_purpose, document_title, due_date])


        session = db.session
        new_unit = session.query(Unit).filter(Unit.code == unit_code).first()
        if new_unit is None:
            new_unit = Unit()
            new_unit.code = unit_code
            session.add(new_unit)
        
        new_discipline = session.query(Discipline).filter(Discipline.name == discipline_name).first()
        if new_discipline is None:
            new_discipline = Discipline()
            new_discipline.code = discipline_code
            new_discipline.name = discipline_name
            session.add(new_discipline)
        
        new_doctype = session.query(Doctype).filter(Doctype.code == doctype).first()
        if new_doctype is None:
            new_doctype = Doctype()
            new_doctype.code = doctype
            session.add(new_doctype)

        
        request = Request()
        request.discipline = new_discipline
        request.unit = new_unit
        request.doctype = new_doctype
        request.quantity = 1
        request.created_by_fk = '1'
        request.changed_by_fk = '1'
        session.add(request)
        session.flush()

        
        new_code = askcode(new_unit, new_discipline,new_doctype,1)
        
        code = Codes()
        code.code = new_code[0]
        code.contractor_code = contractor_code
        code.request_id = request.id
        code.internal_note = note
        code.changed_by_fk = '1'
        code.created_by_fk = '1'
        session.add(code)
        
    session.commit() 